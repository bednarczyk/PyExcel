import openpyxl

wb = openpyxl.load_workbook("workbook.xlsx")
print(wb.sheetnames)
sheet = wb["Sheet1"]
cell = sheet["a1"]
print(cell.value)
print(cell.row)
print(cell.column)
print(cell.coordinate)
cell = sheet.cell(row=1, column=1)
max_row = sheet.max_row
max_col = sheet.max_column
print(max_row, max_col)

for row in range(1, max_row + 1):
    for column in range(1, max_col + 1):
        cell = sheet.cell(row, column)
        print(cell.value)

sheet.append([1, 2, 3])
wb.save("wb2.xlsx")